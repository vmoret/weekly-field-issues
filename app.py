from datetime import datetime as dt
from functools import partial
from itertools import groupby
import os

from openpyxl.workbook import Workbook
import yaml
try:
    from yaml import CLoader as Loader
except ImportError:
    from yaml import Loader

from services import ITrackService
from sprint import calc_handled_in_active_sprint


def load_options(fname, paths):
    fps = (os.path.join(path, fname) for path in paths)
    fp = next(fp for fp in fps if os.path.exists(fp))
    with open(fp, 'r') as stream:
        return yaml.load(stream, Loader=Loader)


def build_jql(options, keys):
    return ' and '.join(options[k] for k in keys)


def calc_severity_counts(itrack, options, *args):
    severities = ('S1 - Major', 'S2 - High', 'S3 - Medium', 'S4 - Low')
    jql = build_jql(options, args) + ' and severity = "{}"'
    counts = [itrack.count(jql.format(x)) for x in severities]
    return [*counts, sum(counts[:2]), sum(counts)]


def calc_created_this_week(itrack, options):
    return calc_severity_counts(itrack, options, 'new')


def add_title(ws, row, title):
    _ = ws.cell(column=1, row=row, value=title)
    return row


def add_measure(ws, row, label, value):
    _ = ws.cell(column=2, row=row, value=label)
    return ws.cell(column=3, row=row, value=value)


def combine(values, keys):
    return zip(range(len(values)), keys, values)


def add_created_this_week(ws, itrack, row, options):
    _ = add_title(ws, row, 'New')
    for n, title, value in combine(calc_created_this_week(itrack, options)[-2:], ('S1+S2', 'Total')):
        _ = add_measure(ws, row + n, title, value)


def add_backlog(ws, itrack, row, options, severities, families):
    _ = add_title(ws, row, 'Active')
    counts = calc_severity_counts(itrack, options, 'Active')
    for n, key, measure in combine(counts, severities):
        _ = add_measure(ws, row + n, key, measure)

    row_1 = add_title(ws, row + 7, 'Unresolved')
    counts = calc_severity_counts(itrack, options, 'Unresolved')
    for n, key, measure in combine(counts, severities):
        _ = add_measure(ws, row_1 + n, key, measure)

    row_2 = add_title(ws, row_1 + 10, 'Unresolved (S1+S2)')
    for n, key in enumerate(families):
        opts = options[key]
        project_jql = 'project in ({})'.format(', '.join('"{}"'.format(x) for x in opts['projects']))
        jql = options['Unresolved (S1+S2)'] + ' and ' + project_jql
        _ = add_measure(ws, row_2 + n, key, itrack.count(jql))


def add_handled_in_active_sprints(ws, itrack, row, options, families,
                                  observed_during, severities):
    _ = add_title(ws, row, 'Active Sprints')
    kwargs = dict(observed_during=observed_during, severities=severities)
    for n, key in enumerate(families):
        opts = options[key]
        if opts['boards']:
            measures = calc_handled_in_active_sprint(itrack, **opts, **kwargs)
        else:
            measures = [-1]
        _ = add_measure(ws, row + n, key, sum(measures))


def main():
    options = load_options('app.yaml', ['/data', os.getcwd()])

    # initialize severities and product families enumerations
    enum = lambda *args: list(enumerate(args))
    severities = ('S1 - Major', 'S2 - High', 'S3 - Medium', 'S4 - Low')
    severities_1 = ('S1', 'S2', 'S3', 'S4', 'S1+S2', 'Total')
    families = ('ClickShare', 'OpSpace', 'TFN', 'SDP', 'RPC',
                'LCD', 'WePresent', 'WeConnect', 'Other')
    observed_during = ('Customer installation', 'Customer use', 'Demo',
                       'Internal Review, FAT, FAI', 'Repair / Service')

    # get auth from environment variables
    auth = tuple(os.getenv(x) for x in ('USER', 'PASSWORD'))
    
    # initialize iTrack ReST API service
    fields = ('key', 'project', 'issuetype', 'customfield_10021',
              'customfield_10002', 'issuelinks')
    itrack = ITrackService(auth=auth, fields=fields)

    # initialize XL workbook
    wb = Workbook()

    # select active sheet and rename to "Data"
    ws = wb.active
    ws.title = 'Data'

    # add title
    ws.cell(row=1, column=1, value='Date')
    ws.cell(row=1, column=3, value=dt.now().date())

    # calc and add field issues created this week
    add_created_this_week(ws, itrack, 16, options)

    # calcl and add field issues backlog
    add_backlog(ws, itrack, 2, options, severities_1, families)

    # calc and add handled in active sprints
    add_handled_in_active_sprints(ws, itrack, 29, options, families, 
                                  observed_during, severities)

    # save XL workbook
    wb.save('/data/weekly-field-issues-{:%Y-%m-%d}.xlsx'.format(dt.now()))
    
if __name__ == '__main__':
    main()

